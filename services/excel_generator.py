import openpyxl
import math
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_takeoff_excel(recon_results: dict, output_path: str, project_name: str, project_type: str = "Single Dwelling") -> None:
    """
    Generates a professionally formatted Excel takeoff sheet
    using openpyxl with multiple tabs, custom color styling, and frozen panes.
    """
    rows = recon_results.get("rows", [])
    flags = recon_results.get("flags", [])
    overall_confidence = recon_results.get("overall_confidence", 0.0)
    is_rejected = recon_results.get("is_rejected", False)
    rejection_reason = recon_results.get("rejection_reason", "")
    review_required = recon_results.get("review_required", False)
    review_reason = recon_results.get("review_reason", "")
    
    wb = openpyxl.Workbook()
    
    # -------------------------------------------------------------
    # TAB 1: CONSISTENCY REPORT
    # -------------------------------------------------------------
    ws_report = wb.active
    ws_report.title = "Consistency Report"
    ws_report.views.sheetView[0].showGridLines = True
    
    # Stylized fonts and fills
    font_title = Font(name="Segoe UI", size=16, bold=True, color="1A2530")
    font_section = Font(name="Segoe UI", size=12, bold=True, color="2F3640")
    font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    font_bold = Font(name="Segoe UI", size=11, bold=True)
    font_regular = Font(name="Segoe UI", size=10)
    font_warning = Font(name="Segoe UI", size=10, bold=True, color="C0392B")
    
    fill_header = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    fill_success = PatternFill(start_color="E8F8F5", end_color="E8F8F5", fill_type="solid")
    fill_danger = PatternFill(start_color="FDEDEC", end_color="FDEDEC", fill_type="solid")
    fill_warning = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid")
    fill_zebra = PatternFill(start_color="F8F9F9", end_color="F8F9F9", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='BDC3C7'),
        right=Side(style='thin', color='BDC3C7'),
        top=Side(style='thin', color='BDC3C7'),
        bottom=Side(style='thin', color='BDC3C7')
    )
    
    double_bottom_border = Border(
        bottom=Side(style='double', color='2C3E50'),
        top=Side(style='thin', color='BDC3C7')
    )
    
    # Title
    ws_report["A1"] = "FenX Takeoff Tool — Consistency & Quality Report"
    ws_report["A1"].font = font_title
    ws_report.row_dimensions[1].height = 30
    
    ws_report["A2"] = f"Project: {project_name} | Type: {project_type}"
    ws_report["A2"].font = Font(name="Segoe UI", size=10, italic=True)
    
    # Summary Card Grid
    ws_report["A4"] = "Takeoff Summary"
    ws_report["A4"].font = font_section
    
    # Summary Table
    summary_headers = ["Metric", "Value"]
    for col_idx, text in enumerate(summary_headers, start=1):
        cell = ws_report.cell(row=5, column=col_idx, value=text)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center")
    
    metrics = [
        ("Total Openings Detected", len(rows)),
        ("Total Consistency Flags Raised", len(flags)),
        ("Overall Confidence Score", f"{overall_confidence:.1f}%"),
        ("Takeoff Process Status",
         "REJECTED (Low Quality)" if is_rejected
         else ("REVIEW REQUIRED" if review_required else "PASSED"))
    ]
    
    for row_offset, (m_name, m_val) in enumerate(metrics, start=6):
        cell_name = ws_report.cell(row=row_offset, column=1, value=m_name)
        cell_val = ws_report.cell(row=row_offset, column=2, value=m_val)
        
        cell_name.font = font_regular
        cell_name.border = thin_border
        cell_val.font = font_bold
        cell_val.border = thin_border
        cell_val.alignment = Alignment(horizontal="center")
        
        # Color highlight code status
        if m_name == "Overall Confidence Score":
            val_num = float(overall_confidence)
            if val_num >= 85:
                cell_val.fill = fill_success
            elif val_num >= 70:
                cell_val.fill = fill_warning
            else:
                cell_val.fill = fill_danger
        elif m_name == "Takeoff Process Status":
            if is_rejected:
                cell_val.fill = fill_danger
                cell_val.font = font_warning
            elif review_required:
                cell_val.fill = fill_warning
                cell_val.font = Font(name="Segoe UI", size=11, bold=True, color="E67E22")
            else:
                cell_val.fill = fill_success
                cell_val.font = Font(name="Segoe UI", size=11, bold=True, color="27AE60")
                
    ws_report.column_dimensions['A'].width = 35
    ws_report.column_dimensions['B'].width = 30
    
    # Flags Table
    start_flags_row = 12
    ws_report.cell(row=start_flags_row, column=1, value="Consistency Flags and Discrepancies").font = font_section
    
    flag_headers = ["Item Ref", "Flag Category", "Description / Resolution Action"]
    for col_idx, text in enumerate(flag_headers, start=1):
        cell = ws_report.cell(row=start_flags_row+1, column=col_idx, value=text)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center" if col_idx < 3 else "left")
        
    ws_report.row_dimensions[start_flags_row+1].height = 25
    
    if not flags:
        ws_report.cell(row=start_flags_row+2, column=1, value="No discrepancies or flags detected. Perfect consistency!").font = Font(name="Segoe UI", size=10, italic=True)
        ws_report.merge_cells(start_row=start_flags_row+2, start_column=1, end_row=start_flags_row+2, end_column=3)
    else:
        for idx, flag in enumerate(flags):
            row_idx = start_flags_row + 2 + idx
            ws_report.row_dimensions[row_idx].height = 20

            # FIX #8: Populate Item Ref and Flag Category from the flag dict
            item_ref = flag.get("item_ref") or flag.get("opening_id") or ""
            category = flag.get("category") or flag.get("flag_type", "").replace("_", " ").title()

            c1 = ws_report.cell(row=row_idx, column=1, value=item_ref)
            c2 = ws_report.cell(row=row_idx, column=2, value=category)
            c3 = ws_report.cell(row=row_idx, column=3, value=flag.get("description", ""))
            
            c1.font = font_bold
            c1.alignment = Alignment(horizontal="center")
            c1.border = thin_border
            
            c2.font = font_bold
            c2.fill = fill_warning if "mismatch" in category.lower() else fill_danger
            c2.alignment = Alignment(horizontal="center")
            c2.border = thin_border
            
            c3.font = font_regular
            c3.border = thin_border
            
            if is_rejected or flag.get("severity") == "High":
                c3.fill = fill_danger
                 
    # Adjust widths for flags columns
    ws_report.column_dimensions['C'].width = 80
    
    # -------------------------------------------------------------
    # TAB 2+: TAKEOFF SHEET(S)
    # -------------------------------------------------------------
    # Setup Dwyer sheets based on project type
    sheet_names = ["Takeoff"]
    if project_type == "Duplex":
        sheet_names = ["Dwelling 1", "Dwelling 2"]
    elif project_type == "Multi-unit":
        # Group by dwellings. In a prototype we can partition them or just output standard sheets.
        sheet_names = ["Unit 1", "Unit 2"]
        
    for name in sheet_names:
        ws = wb.create_sheet(title=name)
        ws.views.sheetView[0].showGridLines = True
        
        # Frozen top rows: Row 1 has metadata, Row 2 is headers, Row 3 is data, freeze row 3 and down
        ws.freeze_panes = ws['A3']
        
        # Metadata row
        ws["A1"] = f"{project_name} — {name} Takeoff schedule"
        ws["A1"].font = Font(name="Segoe UI", size=12, bold=True, color="2C3E50")
        ws.row_dimensions[1].height = 24
        
        # Table Headers
        headers = [
            "Location", "W/D Number", "Height (mm)", "Width (mm)", "Type",
            "Orientation", "Glazing Type", "U-value", "SHGC", "Frame Material",
            "Quantity", "Source Reference"
        ]
        
        for col_idx, text in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col_idx, value=text)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
        ws.row_dimensions[2].height = 28
        
        # Opening type soft pastel colors
        fill_window = PatternFill(start_color="F2F7FD", end_color="F2F7FD", fill_type="solid") # soft pastel blue
        fill_door = PatternFill(start_color="F2FDF8", end_color="F2FDF8", fill_type="solid") # soft pastel green
        fill_bifold = PatternFill(start_color="FDF7F2", end_color="FDF7F2", fill_type="solid") # soft pastel orange
        fill_louvre = PatternFill(start_color="FAF2FD", end_color="FAF2FD", fill_type="solid") # soft pastel purple
        fill_low_conf = PatternFill(start_color="FDECEF", end_color="FDECEF", fill_type="solid") # soft warnings red/pink
        
        # Populate data rows
        # For prototype, we place all rows on each tab, or split them in half if duplex.
        tab_rows = rows
        if project_type == "Duplex" and len(rows) > 1:
            mid = math.ceil(len(rows) / 2)
            if name == "Dwelling 1":
                tab_rows = rows[:mid]
            else:
                tab_rows = rows[mid:]
                
        for row_idx, r in enumerate(tab_rows, start=3):
            ws.row_dimensions[row_idx].height = 20
            
            c1 = ws.cell(row=row_idx, column=1, value=r.get("location", ""))
            c2 = ws.cell(row=row_idx, column=2, value=r.get("tag", ""))
            c3 = ws.cell(row=row_idx, column=3, value=r.get("height", ""))
            c4 = ws.cell(row=row_idx, column=4, value=r.get("width", ""))
            c5 = ws.cell(row=row_idx, column=5, value=r.get("type", ""))
            c6 = ws.cell(row=row_idx, column=6, value=r.get("orientation", ""))
            c7 = ws.cell(row=row_idx, column=7, value=r.get("glazing", ""))
            c8 = ws.cell(row=row_idx, column=8, value=r.get("u_value", "N/A"))
            c9 = ws.cell(row=row_idx, column=9, value=r.get("shgc", "N/A"))
            c10 = ws.cell(row=row_idx, column=10, value=r.get("frame", "Aluminium"))
            c11 = ws.cell(row=row_idx, column=11, value=r.get("quantity", 1))
            c12 = ws.cell(row=row_idx, column=12, value=r.get("src_ref", ""))
            
            # Formatting
            c1.alignment = Alignment(horizontal="left", vertical="center")
            c2.alignment = Alignment(horizontal="center", vertical="center")
            c3.alignment = Alignment(horizontal="right", vertical="center")
            c4.alignment = Alignment(horizontal="right", vertical="center")
            c5.alignment = Alignment(horizontal="left", vertical="center")
            c6.alignment = Alignment(horizontal="center", vertical="center")
            c7.alignment = Alignment(horizontal="left", vertical="center")
            c8.alignment = Alignment(horizontal="center", vertical="center")
            c9.alignment = Alignment(horizontal="center", vertical="center")
            c10.alignment = Alignment(horizontal="left", vertical="center")
            c11.alignment = Alignment(horizontal="right", vertical="center")
            c12.alignment = Alignment(horizontal="left", vertical="center")
            
            # Check row fill color based on opening type and confidence
            opening_type = r.get("opening_type", "Window")
            conf = r.get("confidence", 100.0)
            
            row_fill = fill_window
            if conf < 70.0:
                row_fill = fill_low_conf
            elif opening_type == "Door":
                row_fill = fill_door
            elif opening_type == "Bi-fold/Stacker Door":
                row_fill = fill_bifold
            elif opening_type == "Louvre":
                row_fill = fill_louvre
                
            for cell in [c1, c2, c3, c4, c5, c6, c7, c8, c9, c10, c11, c12]:
                cell.font = font_regular
                cell.fill = row_fill
                cell.border = thin_border
                
        # Auto-fit columns
        for col in ws.columns:
            max_len = 0
            for cell in col:
                if cell.row == 1:
                    continue
                if cell.value:
                    val_str = str(cell.value)
                    if len(val_str) > max_len:
                        max_len = len(val_str)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
    # Save Workbook
    wb.save(output_path)
    wb.close()
